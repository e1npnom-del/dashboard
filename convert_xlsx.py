"""
convert_xlsx.py — แปลง event_outage_data.xlsx → event_outage_data.csv
โครงสร้าง Excel:
  - 1 seq = 1 วัน (merged หลาย row)
  - แต่ละวันมีหลายเหตุการณ์ = sub-rows ที่มี col[8]=สถานที่
  - บาง row col[4]=เวลาเริ่มเป็น None → คำนวณย้อนจาก เวลาสิ้นสุด - ระยะเวลา
  - เหตุการณ์จริง = มีสถานที่ (col[8]) เป็นเงื่อนไขหลัก
"""
import csv, os, re
from datetime import datetime, timedelta
from openpyxl import load_workbook

THAI_MONTHS = {
    'ม.ค.':'01','ก.พ.':'02','มี.ค.':'03','เม.ย.':'04',
    'พ.ค.':'05','มิ.ย.':'06','ก.ค.':'07','ส.ค.':'08',
    'ก.ย.':'09','ต.ค.':'10','พ.ย.':'11','ธ.ค.':'12'
}
THAI_DIGITS = str.maketrans('๐๑๒๓๔๕๖๗๘๙', '0123456789')

HEADERS = [
    'ลำดับ','วันที่','เวลาเริ่ม','เวลาสิ้นสุด','ระยะเวลาดับ_นาที',
    'สถานที่','รหัสอุปกรณ์','จำนวนผชฟ_กระทบ','สภาพอากาศ','สาเหตุ',
    'ช่องทางการแจ้ง','เวลารับแจ้ง','เวลาออกปฏิบัติงาน','เวลาเสร็จงาน',
    'เวลาแก้ไข_นาที','ผู้ปฏิบัติงาน','ยานพาหนะ',
    'อุปกรณ์_รายการ','อุปกรณ์_จำนวน','หมายเหตุ','date_iso'
]

CSV_PATH  = 'event_outage_data.csv'
XLSX_PATH = 'event_outage_data.xlsx'

# ── helpers ──────────────────────────────────────────────────────
def td(s): return str(s).translate(THAI_DIGITS)

def cv(v):
    if v is None: return ''
    if hasattr(v, 'strftime'): return v.strftime('%H:%M')
    s = td(str(v).strip())
    return '' if s in ('None','nan','-') else s

def to_int(v, d=0):
    if v is None: return d
    if hasattr(v, 'strftime'): return d
    try: return max(0, int(float(td(str(v)).strip())))
    except: return d

def to_int_signed(v, d=0):
    """รองรับค่าติดลบ (เช่น ระยะเวลาข้ามเที่ยงคืน)"""
    if v is None: return d
    if hasattr(v, 'strftime'): return d
    try: return int(float(td(str(v)).strip()))
    except: return d

def parse_date(s):
    s = td(str(s).strip())
    for th, num in THAI_MONTHS.items():
        s = s.replace(th, num)
    parts = s.split()
    if len(parts) == 3:
        d, m, y = parts
        try: return f'{int(y)-543}-{m}-{int(d):02d}'
        except: pass
    if re.match(r'^\d{4}-\d{2}-\d{2}$', s): return s
    return ''

def calc_start_time(time_end_str, dur_min):
    """คำนวณเวลาเริ่มย้อนกลับจาก เวลาสิ้นสุด - ระยะเวลา(นาที)"""
    if not time_end_str or dur_min <= 0: return ''
    try:
        m = re.match(r'(\d{1,2}):(\d{2})', time_end_str)
        if not m: return ''
        end_dt = datetime(2000,1,1, int(m.group(1)), int(m.group(2)))
        start_dt = end_dt - timedelta(minutes=dur_min)
        return start_dt.strftime('%H:%M')
    except: return ''

def dedup_key(r):
    return (str(r[20]), str(r[2]), str(r[5]), str(r[6]))

# ── 1. โหลด CSV เก่า ─────────────────────────────────────────────
existing = []
if os.path.exists(CSV_PATH):
    with open(CSV_PATH, newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        next(reader, None)
        for row in reader:
            if row: existing.append(row)
    print(f'📂 โหลดข้อมูลเก่า: {len(existing)} records')
else:
    print(f'📂 ไม่พบ {CSV_PATH} — สร้างใหม่')

existing_keys = {dedup_key(r) for r in existing}

# ── 2. อ่าน xlsx ─────────────────────────────────────────────────
wb = load_workbook(XLSX_PATH, data_only=True)
new_rows = []

for ws in wb.worksheets:
    sheet_new      = 0
    sheet_dup      = 0
    sheet_calc     = 0  # นับ rows ที่คำนวณเวลาเริ่มย้อนกลับ

    # carry-forward values
    last_date_raw = ''
    last_weather  = ''
    last_cause    = ''
    last_channel  = ''
    last_recv     = ''
    last_depart   = ''
    last_finish   = ''
    last_fix      = 0
    last_person   = ''
    last_vehicle  = ''

    for row in ws.iter_rows(min_row=9, values_only=True):
        if len(row) < 10: continue

        # carry-forward วันที่
        if row[2] is not None: last_date_raw = cv(row[2])

        # carry-forward ค่า merged
        if cv(row[11]): last_weather = cv(row[11])
        if cv(row[12]): last_cause   = cv(row[12])
        if cv(row[14]): last_channel = cv(row[14])
        if cv(row[15]): last_recv    = cv(row[15])
        if cv(row[17]): last_depart  = cv(row[17])
        if cv(row[18]): last_finish  = cv(row[18])
        if row[19] is not None: last_fix = to_int(row[19])
        if cv(row[20]): last_person  = cv(row[20]).replace('\n',' | ')
        if cv(row[21]): last_vehicle = cv(row[21])

        # เงื่อนไขหลัก: ต้องมีสถานที่
        loc = cv(row[8])
        if not loc: continue

        # เวลาเริ่ม: ใช้ col[4] ถ้ามี หรือคำนวณย้อนกลับ
        time_start = cv(row[4])
        time_end   = cv(row[6])
        dur        = to_int_signed(row[7])

        if not time_start:
            time_start = calc_start_time(time_end, dur)
            if time_start: sheet_calc += 1

        date_iso = parse_date(last_date_raw)

        r = [
            0,                  # ลำดับ (นับใหม่ทีหลัง)
            last_date_raw,      # วันที่
            time_start,         # เวลาเริ่ม
            time_end,           # เวลาสิ้นสุด
            max(0, dur),        # ระยะเวลาดับ_นาที
            loc,                # สถานที่
            cv(row[9]),         # รหัสอุปกรณ์
            to_int(row[10]),    # จำนวนผชฟ_กระทบ
            last_weather,       # สภาพอากาศ
            last_cause,         # สาเหตุ
            last_channel,       # ช่องทางการแจ้ง
            last_recv,          # เวลารับแจ้ง
            last_depart,        # เวลาออกปฏิบัติงาน
            last_finish,        # เวลาเสร็จงาน
            last_fix,           # เวลาแก้ไข_นาที
            last_person,        # ผู้ปฏิบัติงาน
            last_vehicle,       # ยานพาหนะ
            cv(row[24]) if len(row) > 24 else '',
            cv(row[25]) if len(row) > 25 else '',
            cv(row[29]) if len(row) > 29 else '',
            date_iso,
        ]

        key = dedup_key(r)
        if key in existing_keys:
            sheet_dup += 1
            continue

        new_rows.append(r)
        existing_keys.add(key)
        sheet_new += 1

    print(f'  📄 Sheet "{ws.title}": ใหม่ {sheet_new} | ซ้ำ {sheet_dup} | คำนวณเวลาเริ่มย้อนกลับ {sheet_calc} rows')

print(f'➕ ข้อมูลใหม่: {len(new_rows)} records')

# ── 3. รวม + เรียง + นับลำดับใหม่ ────────────────────────────────
all_rows = existing + new_rows
all_rows.sort(key=lambda r: (str(r[20]), str(r[2])))
for i, r in enumerate(all_rows, 1):
    r[0] = i

# ── 4. เขียน CSV ──────────────────────────────────────────────────
with open(CSV_PATH, 'w', newline='', encoding='utf-8-sig') as f:
    w = csv.writer(f)
    w.writerow(HEADERS)
    w.writerows(all_rows)

print(f'\n✅ บันทึกสำเร็จ: {len(all_rows)} records → {CSV_PATH}')
if all_rows:
    dates = [r[20] for r in all_rows if r[20]]
    if dates: print(f'   ช่วงข้อมูล: {min(dates)} ถึง {max(dates)}')
